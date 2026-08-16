from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import load_workbook

BASE = Path("data/mbie/edgs2024")
ASSUMPTIONS = BASE / "edgs_2024_assumptions.xlsx"
RESULTS = BASE / "edgs_2024_results.xlsx"
MODEL_DIR = BASE / "model"

SCENARIOS = ["Constraint", "Environmental", "Growth", "Innovation", "Reference"]

ASSUMPTION_SHEETS = {
    "datacentres": "Datacentres",
    "ev_share_vkt": "Electric vehicle share of VKT",
    "vehicle_km": "Vehicle kilometres travelled",
    "distributed_solar_pv": "Distributed solar PV",
    "population": "Population",
    "gdp": "GDP",
    "commodity_prices": "Commodity prices",
    "peak_ratios": "Peak ratios",
    "island_share_demand": "Island share of demand",
}

RESULT_SHEETS = {
    "total_electricity_demand": "Total electricity demand",
    "electricity_demand_by_sector": "Electricity demand by sector",
    "peak_demand": "Peak demand",
    "total_energy_demand": "Total energy demand",
}


def sheet_rows(path: Path, sheet_name: str) -> list[dict[str, object]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Missing sheet {sheet_name!r} in {path}")
    ws = wb[sheet_name]
    values = list(ws.iter_rows(values_only=True))
    wb.close()

    header_idx = next(
        (i for i, row in enumerate(values) if row and any(str(v).strip() == "TimePeriod" for v in row if v is not None)),
        None,
    )
    if header_idx is None:
        raise RuntimeError(f"Could not find TimePeriod header in {sheet_name}")

    headers = [str(v).strip() if v is not None else "" for v in values[header_idx]]
    rows: list[dict[str, object]] = []
    for raw in values[header_idx + 1 :]:
        if not raw or all(v is None for v in raw):
            continue
        row = {headers[i]: raw[i] if i < len(raw) else None for i in range(len(headers)) if headers[i]}
        if row.get("TimePeriod") is None:
            continue
        rows.append(row)
    return rows


def write_csv(name: str, rows: list[dict[str, object]]) -> None:
    if not rows:
        return
    MODEL_DIR.mkdir(parents=True, exist_ok=True)
    path = MODEL_DIR / f"{name}.csv"
    fieldnames: list[str] = []
    for row in rows:
        for key in row:
            if key not in fieldnames:
                fieldnames.append(key)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    print(f"Wrote {path} ({len(rows)} rows)")


def projected_summary(
    rows: list[dict[str, object]],
    years=(2024, 2030, 2040, 2050),
    *,
    variable: str | None = None,
    unit: str | None = None,
) -> dict[str, dict[str, float]]:
    out: dict[str, dict[str, float]] = {}
    for scenario in SCENARIOS:
        year_map: dict[str, float] = {}
        for row in rows:
            if str(row.get("Scenario")) != scenario:
                continue
            if variable is not None and str(row.get("Variable")) != variable:
                continue
            if unit is not None and str(row.get("Unit")) != unit:
                continue
            try:
                year = int(row["TimePeriod"])
            except (TypeError, ValueError):
                continue
            if year not in years:
                continue
            value = row.get("Value")
            if isinstance(value, (int, float)):
                year_map[str(year)] = float(value)
        out[scenario] = year_map
    return out


def sector_summary(rows: list[dict[str, object]], years=(2024, 2030, 2040, 2050)) -> dict[str, dict[str, dict[str, float]]]:
    out: dict[str, dict[str, dict[str, float]]] = {s: {} for s in SCENARIOS}
    for row in rows:
        scenario = str(row.get("Scenario"))
        if scenario not in out:
            continue
        try:
            year = int(row["TimePeriod"])
        except (TypeError, ValueError):
            continue
        if year not in years:
            continue
        sector = str(row.get("Sector") or row.get("Variable") or "Unknown")
        value = row.get("Value")
        if isinstance(value, (int, float)):
            out[scenario].setdefault(str(year), {})[sector] = float(value)
    return out


def main() -> None:
    normalized: dict[str, list[dict[str, object]]] = {}

    for key, sheet in ASSUMPTION_SHEETS.items():
        rows = sheet_rows(ASSUMPTIONS, sheet)
        normalized[key] = rows
        write_csv(key, rows)

    for key, sheet in RESULT_SHEETS.items():
        rows = sheet_rows(RESULTS, sheet)
        normalized[key] = rows
        write_csv(key, rows)

    summary = {
        "source": "MBIE Electricity Demand and Generation Scenarios 2024",
        "scenarios": SCENARIOS,
        "selected_years": [2024, 2030, 2040, 2050],
        "total_electricity_demand": projected_summary(normalized["total_electricity_demand"]),
        "electricity_demand_by_sector": sector_summary(normalized["electricity_demand_by_sector"]),
        "peak_demand": projected_summary(normalized["peak_demand"]),
        "datacentre_load_capacity_mw": projected_summary(
            normalized["datacentres"], variable="Load capacity", unit="MW"
        ),
        "datacentre_electricity_consumption_twh": projected_summary(
            normalized["datacentres"], variable="Electricity consumption", unit="TWh"
        ),
        "ev_share_of_vkt": projected_summary(normalized["ev_share_vkt"]),
        "notes": {
            "use": "Use EDGS total electricity demand as the main future-demand trajectory, with sector and assumption sheets retained for diagnostics and sensitivity cases.",
            "thermal": "Thermal generation is treated elsewhere as an available but potentially high-cost security backstop rather than assumed unavailable because of gas scarcity.",
            "btm": "Our reconciled-demand baseline separately adds estimated behind-meter retained residential PV so future demand should be aligned to underlying consumption definitions before replay.",
            "datacentres": "The EDGS Datacentres sheet contains both load capacity (MW) and annual electricity consumption (TWh); these are retained as separate summary fields.",
        },
    }
    MODEL_DIR.mkdir(parents=True, exist_ok=True)
    path = MODEL_DIR / "demand_scenarios_summary.json"
    path.write_text(json.dumps(summary, indent=2, default=str) + "\n", encoding="utf-8")
    print(f"Wrote {path}")


if __name__ == "__main__":
    main()
