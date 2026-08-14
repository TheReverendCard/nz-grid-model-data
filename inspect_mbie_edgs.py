from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

DATA_DIR = Path("data/mbie/edgs2024")
OUTPUT = DATA_DIR / "workbook_structure.json"


def inspect(path: Path) -> dict[str, object]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    for ws in workbook.worksheets:
        sample = []
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True):
            values = [value for value in row]
            if any(value is not None for value in values):
                sample.append(values[:20])
            if len(sample) >= 8:
                break
        sheets.append({
            "name": ws.title,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "sample_nonempty_rows": sample,
        })
    workbook.close()
    return {"file": str(path), "sheets": sheets}


def main() -> None:
    payload = {
        "assumptions": inspect(DATA_DIR / "edgs_2024_assumptions.xlsx"),
        "results": inspect(DATA_DIR / "edgs_2024_results.xlsx"),
    }
    OUTPUT.write_text(json.dumps(payload, indent=2, default=str) + "\n", encoding="utf-8")
    print(f"Wrote {OUTPUT}")
    for book, details in payload.items():
        print(book)
        for sheet in details["sheets"]:
            print(f"  {sheet['name']}: {sheet['max_row']} x {sheet['max_column']}")


if __name__ == "__main__":
    main()
