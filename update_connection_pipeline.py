from __future__ import annotations

import csv
import hashlib
import json
import re
from pathlib import Path

import requests
from openpyxl import load_workbook

URL = "https://static.transpower.co.nz/public/uncontrolled_docs/Generation%20and%20energy%20storage%20connection%20pipeline.xlsx"
DATA_DIR = Path("data/pipeline")
RAW = DATA_DIR / "generation_and_energy_storage_connection_pipeline.xlsx"
OUT = DATA_DIR / "transpower_generation_storage_pipeline.csv"
META = Path("data/metadata/connection_pipeline_source.json")


def load_previous() -> dict[str, object]:
    if not META.exists():
        return {}
    try:
        return json.loads(META.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}


def clean(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


def num(value: object):
    if value is None or value == "":
        return ""
    match = re.search(r"-?\d+(?:\.\d+)?", str(value).replace(",", ""))
    return float(match.group()) if match else ""


def find(headers: list[str], *needles: str) -> int | None:
    for index, header in enumerate(headers):
        if any(needle in clean(header) for needle in needles):
            return index
    return None


def download() -> tuple[dict[str, str], bool]:
    previous = load_previous()
    headers: dict[str, str] = {}
    if RAW.exists():
        etag = str(previous.get("etag") or "")
        last_modified = str(previous.get("last_modified") or "")
        if etag:
            headers["If-None-Match"] = etag
        if last_modified:
            headers["If-Modified-Since"] = last_modified

    response = requests.get(URL, headers=headers, timeout=120)
    if response.status_code == 304 and RAW.exists():
        print(f"Unchanged {RAW} (HTTP 304)")
        return {
            "source_url": str(previous.get("source_url") or URL),
            "etag": str(previous.get("etag") or ""),
            "last_modified": str(previous.get("last_modified") or ""),
        }, False

    response.raise_for_status()
    if not response.content.startswith(b"PK"):
        raise RuntimeError("Transpower pipeline response is not XLSX")

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    changed = not RAW.exists() or RAW.read_bytes() != response.content
    if changed:
        RAW.write_bytes(response.content)
        print(f"Wrote {RAW} ({len(response.content):,} bytes)")
    else:
        print(f"Unchanged {RAW} (byte-identical response)")
    return {
        "source_url": response.url,
        "etag": response.headers.get("ETag", ""),
        "last_modified": response.headers.get("Last-Modified", ""),
    }, changed


def normalize(source: dict[str, str]) -> None:
    workbook = load_workbook(RAW, data_only=True, read_only=True)
    best = None
    for worksheet in workbook.worksheets:
        for row_index, row in enumerate(
            worksheet.iter_rows(min_row=1, max_row=min(worksheet.max_row, 30), values_only=True),
            1,
        ):
            values = [str(value or "") for value in row]
            joined = " ".join(values).lower()
            if (
                len([value for value in values if value.strip()]) >= 4
                and ("project" in joined or "connection" in joined)
                and ("mw" in joined or "capacity" in joined)
            ):
                best = (worksheet, row_index, list(row))
                break
        if best:
            break
    if not best:
        workbook.close()
        raise RuntimeError("Could not identify pipeline header row")

    worksheet, header_row, raw_headers = best
    headers: list[str] = []
    seen: dict[str, int] = {}
    for index, header in enumerate(raw_headers):
        base = clean(header) or f"column_{index + 1}"
        seen[base] = seen.get(base, 0) + 1
        headers.append(base if seen[base] == 1 else f"{base}_{seen[base]}")

    project = find(headers, "project_name", "project")
    technology = find(headers, "technology", "generation_type", "fuel_type", "resource_type", "subtype")
    capacity = find(headers, "capacity_mw", "maximum_capacity", "capacity", "mw")
    stage = find(headers, "stage", "status")
    region = find(headers, "region")
    location = find(headers, "location", "point_of_connection", "poc")
    customer = find(headers, "customer", "developer", "proponent")
    expected_date = find(headers, "need_date", "commission", "connection_date")

    rows: list[dict[str, object]] = []
    for raw_row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
        values = list(raw_row) + [None] * max(0, len(headers) - len(raw_row))
        if not any(value not in (None, "") for value in values):
            continue
        record: dict[str, object] = {
            "project_name": values[project] if project is not None else "",
            "technology": values[technology] if technology is not None else "",
            "capacity_mw": num(values[capacity]) if capacity is not None else "",
            "stage": values[stage] if stage is not None else "",
            "region": values[region] if region is not None else "",
            "location": values[location] if location is not None else "",
            "customer_or_developer": values[customer] if customer is not None else "",
            "expected_connection_or_need_date": values[expected_date] if expected_date is not None else "",
        }
        for header, value in zip(headers, values):
            record["raw_" + header] = value if value is not None else ""
        rows.append(record)

    workbook.close()
    if not rows:
        raise RuntimeError("Transpower pipeline normalization produced no rows")

    with OUT.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)

    payload = {
        "source": "Transpower",
        "source_url": source["source_url"],
        "sha256": hashlib.sha256(RAW.read_bytes()).hexdigest(),
        "etag": source["etag"],
        "last_modified": source["last_modified"],
        "worksheet": worksheet.title,
        "header_row": header_row,
        "rows": len(rows),
    }
    META.parent.mkdir(parents=True, exist_ok=True)
    META.write_text(json.dumps(payload, indent=2, default=str) + "\n", encoding="utf-8")
    print(f"Wrote {OUT} ({len(rows)} rows)")


def main() -> None:
    source, changed = download()
    if not changed and OUT.exists() and META.exists():
        print("Transpower connection pipeline unchanged; skipped workbook normalization")
        return
    normalize(source)


if __name__ == "__main__":
    main()
